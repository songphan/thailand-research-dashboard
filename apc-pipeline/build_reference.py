#!/usr/bin/env python3
"""
Build a unified APC reference table from the publisher price lists in ../apc/.

Validated lists (Elsevier, Wiley x2, Springer x2, Cambridge UP, Oxford UP) use
explicit parsers. Any clean file dropped in ../apc/ is parsed by a generic
auto-detector (header row + ISSN + title + USD/EUR/GBP). Crawled publisher lists
that lack ISSNs live in ../apc/crawled/ as CSVs with the header
"publisher,title,issn,model,usd,eur,gbp" and are matched later by journal title.

Output: apc_reference.json and apc_reference.csv. Fully offline.
"""

import csv
import json
import os
import re

import openpyxl
import pdfplumber

HERE = os.path.dirname(os.path.abspath(__file__))
APC_DIR = os.path.join(HERE, "..", "apc")
CRAWLED_DIR = os.path.join(APC_DIR, "crawled")

EURO = "€"
POUND = "£"
EMDASH = "—"
SPACES = re.compile("[\\s  ]+")

PUBLISHER_KEYWORDS = [
    (r"taylor|francis|t&f|tandf", "Taylor & Francis"),
    (r"\bsage\b", "SAGE"),
    (r"oxford|\boup\b", "Oxford University Press"),
    (r"cambridge|\bcup\b|cupa", "Cambridge University Press"),
    (r"\bacs\b|american chemical", "American Chemical Society"),
    (r"\bieee\b", "IEEE"),
    (r"emerald", "Emerald"),
    (r"\brsc\b|royal society of chem", "Royal Society of Chemistry"),
    (r"iopp|\biop\b|institute of physics", "IOP Publishing"),
    (r"wolters|kluwer|lippincott", "Wolters Kluwer"),
    (r"\bbmj\b", "BMJ"),
    (r"dove", "Dove Press"),
    (r"\baps\b|american physical", "American Physical Society"),
    (r"edp", "EDP Sciences"),
    (r"elsevier|sciencedirect", "Elsevier"),
    (r"wiley", "Wiley"),
    (r"springer|nature", "Springer Nature"),
]

TITLE_STOP = re.compile(r"\b(the|a|an|of|and|for|in|on)\b", re.I)


def norm_issn(raw):
    if raw is None:
        return None
    s = SPACES.sub("", str(raw).strip().upper()).replace("-", "")
    if len(s) == 7:
        s = "0" + s  # integer ISSN that dropped a single leading zero
    if len(s) != 8 or not re.fullmatch(r"\d{7}[\dX]", s):
        return None
    return s[:4] + "-" + s[4:]


def norm_title(t):
    if not t:
        return None
    s = str(t).lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = TITLE_STOP.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def to_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = SPACES.sub("", str(v).strip()).replace(",", "")
    if s == "" or s.lower() in ("n/a", "na", "-", EMDASH, "free", "tbc", "0"):
        return None
    m = re.search(r"\d+(\.\d+)?", s)
    return int(round(float(m.group()))) if m else None


def infer_publisher(filename):
    low = filename.lower()
    for pat, name in PUBLISHER_KEYWORDS:
        if re.search(pat, low):
            return name
    base = re.sub(r"[_\-]+", " ", os.path.splitext(filename)[0])
    base = re.sub(r"\b(apc|apcs|price|list|journals?|charges?|20\d\d|hybrid|open access|fully)\b",
                  "", base, flags=re.I)
    return re.sub(r"\s+", " ", base).strip() or filename


def infer_model(filename):
    low = filename.lower()
    if re.search(r"hybrid|online ?open|transformative", low):
        return "hybrid"
    if re.search(r"fully open|open access|gold|\boa\b", low):
        return "fully_oa"
    return "unknown"


records = []
collisions = 0
seen = {}        # issn_norm -> index
seen_title = {}  # title_norm -> index (for title-keyed records without ISSN)


def add(issn_raw, title, publisher, model, usd, eur, gbp, source, imprint=None):
    global collisions
    issn = norm_issn(issn_raw)
    if not issn:
        return False
    issn_norm = issn.replace("-", "")
    rec = {"issn": issn, "issn_norm": issn_norm,
           "title": (str(title).strip() if title else None),
           "title_norm": norm_title(title),
           "publisher": publisher, "imprint": imprint, "model": model,
           "usd": usd, "eur": eur, "gbp": gbp, "source": source}
    if issn_norm in seen:
        collisions += 1
        prev = records[seen[issn_norm]]
        if prev.get("usd") is None and usd is not None:
            records[seen[issn_norm]] = rec
        return False
    seen[issn_norm] = len(records)
    records.append(rec)
    return True


def add_titlekeyed(title, publisher, model, usd, eur, gbp, source):
    tn = norm_title(title)
    if not tn or (usd is None and eur is None and gbp is None):
        return False
    if tn in seen_title:
        return False
    rec = {"issn": None, "issn_norm": None,
           "title": (str(title).strip() if title else None), "title_norm": tn,
           "publisher": publisher, "imprint": None, "model": model,
           "usd": usd, "eur": eur, "gbp": gbp, "source": source}
    seen_title[tn] = len(records)
    records.append(rec)
    return True


def parse_elsevier(path, filename):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
    model_map = {"fully open access": "fully_oa", "hybrid": "hybrid",
                 "subsidized": "subsidized", "none": "none"}
    n = 0
    for r in ws.iter_rows(min_row=5, values_only=True):
        issn, title, bm, usd, eur, gbp = r[0], r[1], r[2], r[3], r[4], r[5]
        if not issn or not title:
            continue
        model = model_map.get(str(bm).strip().lower(), "none") if bm else "none"
        if add(issn, title, "Elsevier", model, to_int(usd), to_int(eur), to_int(gbp), filename):
            n += 1
    return n


def parse_wiley(path, filename, model, header_row, col_usd, col_gbp, col_eur,
                col_issn=2, col_title=0):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
    n = 0
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        title, issn = r[col_title], r[col_issn]
        if not issn or not title:
            continue
        if add(issn, title, "Wiley", model, to_int(r[col_usd]), to_int(r[col_eur]),
               to_int(r[col_gbp]), filename):
            n += 1
    return n


def parse_springer(path, filename, model):
    n = 0
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 6:
                    continue
                title, imprint, issn, eur, usd, gbp = row[:6]
                if issn and re.search(r"issn", str(issn), re.I):
                    continue
                if not issn or not title:
                    continue
                if add(issn, title, "Springer Nature", model, to_int(usd), to_int(eur),
                       to_int(gbp), filename, imprint=(str(imprint).strip() if imprint else None)):
                    n += 1
    return n


def parse_cup(path, filename):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True).worksheets[0]
    model_map = {"gold oa": "fully_oa", "hybrid": "hybrid"}
    n = 0
    for r in ws.iter_rows(min_row=6, values_only=True):
        if len(r) < 9:
            continue
        issn = r[3] or r[2]
        title = r[1]
        if not issn or not title:
            continue
        model = model_map.get(str(r[7]).strip().lower() if r[7] else "", "unknown")
        if add(issn, title, "Cambridge University Press", model, None, None, to_int(r[8]), filename):
            n += 1
    return n


def parse_oup(path, filename):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True).worksheets[0]
    model_map = {"hybrid": "hybrid", "full open access": "fully_oa"}
    n = 0
    for r in ws.iter_rows(min_row=17, values_only=True):
        if len(r) < 11:
            continue
        issn, title = r[9], r[1]
        if not issn or not title:
            continue
        cur = str(r[6]).strip().upper() if r[6] else ""
        amt = to_int(r[7])
        usd = amt if cur == "USD" else None
        eur = amt if cur == "EUR" else None
        gbp = amt if cur == "GBP" else None
        model = model_map.get(str(r[10]).strip().lower() if r[10] else "", "unknown")
        if add(issn, title, "Oxford University Press", model, usd, eur, gbp, filename):
            n += 1
    return n


def parse_crawled_csv(path, filename):
    # Standard crawled format: publisher,title,issn,model,usd,eur,gbp.
    # Rows with an ISSN are ISSN-keyed; rows without are title-keyed.
    n = 0
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        for row in csv.DictReader(f):
            pub = (row.get("publisher") or "").strip()
            title = (row.get("title") or "").strip()
            model = (row.get("model") or "unknown").strip() or "unknown"
            usd, eur, gbp = to_int(row.get("usd")), to_int(row.get("eur")), to_int(row.get("gbp"))
            issn = row.get("issn")
            if not title or not pub:
                continue
            if norm_issn(issn):
                if add(issn, title, pub, model, usd, eur, gbp, filename):
                    n += 1
            else:
                if add_titlekeyed(title, pub, model, usd, eur, gbp, filename):
                    n += 1
    return n


def is_crawled_csv(path):
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            head = f.readline().strip().lower()
        return head.startswith("publisher,title,issn,model,usd,eur,gbp")
    except Exception:
        return False


ISSN_HDR = re.compile(r"e-?issn|online\s*issn|issn", re.I)
TITLE_HDR = re.compile(r"journal|title|name", re.I)
NOT_TITLE = re.compile(r"subject|area|publisher|imprint|category|discipline", re.I)
USD_HDR = re.compile(r"\busd\b|us\s*\$|\$|dollar", re.I)
EUR_HDR = re.compile(r"\beur\b|euro|" + EURO, re.I)
GBP_HDR = re.compile(r"\bgbp\b|pound|" + POUND, re.I)
MODEL_HDR = re.compile(r"business model|^model$|open access|access type", re.I)
MODEL_VAL = {"fully open access": "fully_oa", "full open access": "fully_oa",
             "open access": "fully_oa", "gold": "fully_oa", "hybrid": "hybrid",
             "subsidized": "subsidized", "none": "none"}


def _detect_columns(grid, max_scan=20):
    for idx in range(min(max_scan, len(grid))):
        cells = [("" if c is None else str(c)) for c in grid[idx]]
        issn_col = None
        for pref in (re.compile(r"e-?issn|online\s*issn", re.I), ISSN_HDR):
            for j, c in enumerate(cells):
                if c and pref.search(c):
                    issn_col = j
                    break
            if issn_col is not None:
                break
        if issn_col is None:
            continue
        colmap = {"issn": issn_col}
        for j, c in enumerate(cells):
            if not c:
                continue
            if "title" not in colmap and TITLE_HDR.search(c) and not NOT_TITLE.search(c):
                colmap["title"] = j
            if "model" not in colmap and MODEL_HDR.search(c):
                colmap["model"] = j
        currency_offset = 0
        for off in (0, 1, 2):
            if idx + off >= len(grid):
                continue
            crow = [("" if c is None else str(c)) for c in grid[idx + off]]
            found = {}
            for j, c in enumerate(crow):
                if c and "usd" not in found and USD_HDR.search(c):
                    found["usd"] = j
                if c and "eur" not in found and EUR_HDR.search(c):
                    found["eur"] = j
                if c and "gbp" not in found and GBP_HDR.search(c):
                    found["gbp"] = j
            if found:
                colmap.update(found)
                currency_offset = off
                break
        if "title" not in colmap:
            colmap["title"] = 0 if issn_col != 0 else 1
        return idx, colmap, currency_offset
    return None


def _xlsx_grid(path):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _csv_grid(path):
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return [row for row in csv.reader(f)]


def parse_generic_tabular(path, filename, publisher, file_model, grid):
    det = _detect_columns(grid)
    report = {"file": filename, "publisher": publisher, "rows": 0, "priced": 0,
              "columns": None, "model_source": None, "sample": [], "warning": None}
    if not det:
        report["warning"] = "Could not detect a header row with an ISSN column."
        return 0, report
    hidx, cm, coff = det
    report["columns"] = cm
    n = 0
    for r in grid[hidx + coff + 1:]:
        if not r or len(r) <= cm["issn"]:
            continue
        issn = r[cm["issn"]]
        title = r[cm["title"]] if cm["title"] < len(r) else None
        if not issn:
            continue
        model = file_model
        if "model" in cm and cm["model"] < len(r) and r[cm["model"]]:
            model = MODEL_VAL.get(str(r[cm["model"]]).strip().lower(), file_model)
            report["model_source"] = "column"
        else:
            report["model_source"] = "filename"
        usd = to_int(r[cm["usd"]]) if "usd" in cm and cm["usd"] < len(r) else None
        eur = to_int(r[cm["eur"]]) if "eur" in cm and cm["eur"] < len(r) else None
        gbp = to_int(r[cm["gbp"]]) if "gbp" in cm and cm["gbp"] < len(r) else None
        if add(issn, title, publisher, model, usd, eur, gbp, filename):
            n += 1
            if usd is not None or gbp is not None or eur is not None:
                report["priced"] += 1
            if len(report["sample"]) < 3:
                report["sample"].append({"issn": norm_issn(issn),
                                         "title": (str(title)[:40] if title else None),
                                         "usd": usd, "gbp": gbp, "eur": eur, "model": model})
    report["rows"] = n
    if "usd" not in cm and "eur" not in cm and "gbp" not in cm:
        report["warning"] = "No currency column detected; prices are empty."
    return n, report


def parse_generic_pdf(path, filename, publisher, file_model):
    report = {"file": filename, "publisher": publisher, "rows": 0, "priced": 0,
              "columns": None, "model_source": "filename", "sample": [], "warning": None}
    cm = None
    coff = 0
    n = 0
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            if cm is None:
                det = _detect_columns(table, max_scan=4)
                if not det:
                    continue
                hidx, cm, coff = det
                report["columns"] = cm
                rows = table[hidx + coff + 1:]
            else:
                rows = [r for r in table if not (r and len(r) > cm["issn"] and r[cm["issn"]]
                        and re.search(r"issn", str(r[cm["issn"]]), re.I))]
            for r in rows:
                if not r or len(r) <= cm["issn"]:
                    continue
                issn = r[cm["issn"]]
                title = r[cm["title"]] if cm["title"] < len(r) else None
                if not issn:
                    continue
                usd = to_int(r[cm["usd"]]) if "usd" in cm and cm["usd"] < len(r) else None
                eur = to_int(r[cm["eur"]]) if "eur" in cm and cm["eur"] < len(r) else None
                gbp = to_int(r[cm["gbp"]]) if "gbp" in cm and cm["gbp"] < len(r) else None
                if add(issn, title, publisher, file_model, usd, eur, gbp, filename):
                    n += 1
                    if usd is not None or gbp is not None or eur is not None:
                        report["priced"] += 1
                    if len(report["sample"]) < 3:
                        report["sample"].append({"issn": norm_issn(issn),
                                                 "title": (str(title)[:40] if title else None),
                                                 "usd": usd})
    if cm is None:
        report["warning"] = "Could not detect a table header with an ISSN column."
    report["rows"] = n
    return n, report


def process_file(path, filename):
    if filename == "CUPA_APC_Pricelist_2026_May.xlsx":
        return parse_cup(path, filename), None
    if filename == "OUP charges.xlsx":
        return parse_oup(path, filename), None
    if filename == "SD_APC.xlsx":
        return parse_elsevier(path, filename), None
    if filename == "Wiley-Journal-APCs-OnlineOpen (1).xlsx":
        return parse_wiley(path, filename, "hybrid", 7, 4, 5, 6), None
    if filename == "Wiley-Journal-APCs-Open-Access (1).xlsx":
        return parse_wiley(path, filename, "fully_oa", 7, 4, 5, 6), None
    if filename == "2026 Springer Nature fully open access journals.pdf":
        return parse_springer(path, filename, "fully_oa"), None
    if filename == "2026 Springer Nature hybrid journals.pdf":
        return parse_springer(path, filename, "hybrid"), None
    publisher = infer_publisher(filename)
    file_model = infer_model(filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return parse_generic_tabular(path, filename, publisher, file_model, _xlsx_grid(path))
    if ext in (".csv", ".tsv"):
        return parse_generic_tabular(path, filename, publisher, file_model, _csv_grid(path))
    if ext == ".pdf":
        return parse_generic_pdf(path, filename, publisher, file_model)
    return 0, {"file": filename, "warning": "Unsupported extension " + ext, "rows": 0}


def main():
    files = sorted(f for f in os.listdir(APC_DIR)
                   if not f.startswith("~") and not f.startswith("."))
    bespoke_counts = {}
    generic_reports = []
    crawled_counts = {}

    for f in files:
        path = os.path.join(APC_DIR, f)
        if not os.path.isfile(path):
            continue
        count, report = process_file(path, f)
        if report is None:
            bespoke_counts[f] = count
        else:
            generic_reports.append(report)

    # Crawled, title-keyed CSVs in ../apc/crawled/
    if os.path.isdir(CRAWLED_DIR):
        for f in sorted(os.listdir(CRAWLED_DIR)):
            path = os.path.join(CRAWLED_DIR, f)
            if os.path.isfile(path) and f.lower().endswith(".csv") and is_crawled_csv(path):
                crawled_counts[f] = parse_crawled_csv(path, "crawled/" + f)

    if not records:
        print("No records parsed. Check the files in ../apc/.")
        return

    with open(os.path.join(HERE, "apc_reference.json"), "w", encoding="utf-8") as fh:
        json.dump(records, fh, ensure_ascii=False, indent=1)
    with open(os.path.join(HERE, "apc_reference.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=list(records[0].keys()))
        w.writeheader()
        w.writerows(records)

    # Compact client pricing file for the dashboard's live APC panel:
    # issn_norm -> [usd, eur, gbp] and title_norm -> [usd, eur, gbp].
    issn_map, title_map = {}, {}
    for r in records:
        val = [r.get("usd"), r.get("eur"), r.get("gbp")]
        if all(v is None for v in val):
            continue
        if r.get("issn_norm") and r["issn_norm"] not in issn_map:
            issn_map[r["issn_norm"]] = val
        if r.get("title_norm") and r["title_norm"] not in title_map:
            title_map[r["title_norm"]] = val
    data_dir = os.path.join(HERE, "..", "src", "data")
    os.makedirs(data_dir, exist_ok=True)
    with open(os.path.join(data_dir, "apc_pricing.json"), "w", encoding="utf-8") as fh:
        json.dump({"issn": issn_map, "title": title_map}, fh, ensure_ascii=False, separators=(",", ":"))

    print("Validated (bespoke) parsers:")
    for k, v in bespoke_counts.items():
        print("  %-48s %d" % (k[:48], v))
    if generic_reports:
        print("\nAuto-detected files (VERIFY THESE):")
        for r in generic_reports:
            print("  " + r["file"])
            print("    publisher: %s   rows: %d   priced: %d" % (r["publisher"], r["rows"], r["priced"]))
            print("    columns:   %s   model from: %s" % (r.get("columns"), r.get("model_source")))
            if r.get("sample"):
                print("    sample:    %s" % r["sample"])
            if r.get("warning"):
                print("    WARNING:   %s" % r["warning"])
    if crawled_counts:
        print("\nCrawled title-keyed CSVs:")
        for k, v in crawled_counts.items():
            print("  %-44s %d rows" % (k, v))

    by_pub, no_issn, by_model = {}, 0, {}
    for r in records:
        by_pub[r["publisher"]] = by_pub.get(r["publisher"], 0) + 1
        by_model[r["model"]] = by_model.get(r["model"], 0) + 1
        if r["issn"] is None:
            no_issn += 1
    print("\nTotal records: %d  (ISSN-keyed: %d, title-keyed: %d)" % (len(records), len(records) - no_issn, no_issn))
    print("Duplicate-ISSN collisions:  %d" % collisions)
    print("By publisher: %s" % by_pub)
    print("By model:     %s" % by_model)


if __name__ == "__main__":
    main()
